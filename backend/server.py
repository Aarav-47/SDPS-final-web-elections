from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Header
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import io
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ---------- Bootstrap ----------
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Accept either MONGO_URL (preferred) or MONGO_URI (Azure Portal naming).
# Falls back to the SDPS Atlas cluster so the app auto-connects on any deployment
# without requiring env-var configuration. Override via Azure App Service settings
# if you ever rotate the credentials.
MONGO_URL = (
    os.environ.get('MONGO_URL')
    or os.environ.get('MONGO_URI')
    or 'mongodb+srv://SDPS:SDPS@sdps-election-server.trpp58b.mongodb.net/?appName=SDPS-Election-Server'
)

DB_NAME = os.environ.get('DB_NAME', 'sdps-election')
JWT_SECRET = os.environ.get('JWT_SECRET', 'sdps-election-prod-secret-Krish2026-rotate-me')
JWT_ALGO = 'HS256'

# Async Mongo client (motor) — pooled, used for all queries.
# retryWrites=True is the Atlas default and works out of the box.
# (For Cosmos DB Mongo API, append &retrywrites=false to the URL itself.)
client = AsyncIOMotorClient(
    MONGO_URL,
    serverSelectionTimeoutMS=10000,
    connectTimeoutMS=10000,
    maxPoolSize=50,
)
db = client[DB_NAME]

DEFAULT_POSTS = [
    {"key": "head_boy", "title": "Head Boy", "order": 1},
    {"key": "head_girl", "title": "Head Girl", "order": 2},
    {"key": "sports_skipper", "title": "Sports Skipper", "order": 3},
    {"key": "cultural_head", "title": "Cultural Head", "order": 4},
    {"key": "discipline_head", "title": "Discipline Head", "order": 5},
]

app = FastAPI(title="SDPS Student Council Election")
api = APIRouter(prefix="/api")


# ---------- Models ----------
class User(BaseModel):
    admission_no: str
    name: str
    role: str = "student"
    father_name: str = ""
    class_name: str = ""
    subject: str = ""
    designation: str = ""

class Candidate(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    post: str
    name: str
    photo: str = ""
    symbol: str = ""
    symbol_image: str = ""
    adjustment: int = 0

class CandidateCreate(BaseModel):
    post: str
    name: str
    photo: str = ""
    symbol: str = ""
    symbol_image: str = ""
    adjustment: int = 0

class CandidateUpdate(BaseModel):
    name: Optional[str] = None
    photo: Optional[str] = None
    symbol: Optional[str] = None
    symbol_image: Optional[str] = None
    post: Optional[str] = None
    adjustment: Optional[int] = None

class PostCreate(BaseModel):
    title: str
    order: Optional[int] = None

class PostUpdate(BaseModel):
    title: Optional[str] = None
    order: Optional[int] = None

class Ballot(BaseModel):
    admission_no: str
    selections: Dict[str, str]

class BallotUpdate(BaseModel):
    selections: Dict[str, str]

class AdminLogin(BaseModel):
    username: str
    password: str

class SettingValue(BaseModel):
    value: str


# ---------- Helpers ----------
def make_token(username: str) -> str:
    payload = {"sub": username, "exp": datetime.now(timezone.utc) + timedelta(hours=12)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)

def verify_admin(authorization: Optional[str] = Header(None)) -> str:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing token")
    token = authorization.split(" ", 1)[1]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
        return payload["sub"]
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

def slugify(s: str) -> str:
    base = "".join(c.lower() if c.isalnum() else "_" for c in s).strip("_")
    while "__" in base:
        base = base.replace("__", "_")
    return base or "post"

async def active_post_keys() -> List[str]:
    docs = await db.posts.find({}, {"_id": 0}).sort("order", 1).to_list(1000)
    return [d["key"] for d in docs]


PUBLIC_BACKEND_URL = os.environ.get(
    'PUBLIC_BACKEND_URL',
    'https://sdps-election-rg-d9cqbwakd4exb8d0.centralindia-01.azurewebsites.net',
).rstrip('/')


def _lighten_candidate(c: Dict) -> Dict:
    """Replace heavy base64 data-URI photos with a lazy-loaded URL.
    Keeps small HTTPS photo URLs as-is. Critical for /candidates and /bootstrap
    response size on Azure (was 12 MB → ~5 KB)."""
    out = dict(c)
    photo = out.get("photo") or ""
    if photo.startswith("data:"):
        out["photo"] = f"{PUBLIC_BACKEND_URL}/api/candidates/{c['id']}/photo"
    sym = out.get("symbol_image") or ""
    if sym.startswith("data:"):
        out["symbol_image"] = f"{PUBLIC_BACKEND_URL}/api/candidates/{c['id']}/symbol"
    return out


# ---------- Seeding & Indexes ----------
async def ensure_indexes():
    """Critical for performance: indexes on frequently queried fields."""
    try:
        await db.users.create_index("admission_no", unique=True)
        await db.votes.create_index("admission_no", unique=True)
        await db.candidates.create_index("post")
        await db.candidates.create_index("id", unique=True)
        await db.posts.create_index("key", unique=True)
        await db.posts.create_index("order")
        await db.admins.create_index("username", unique=True)
        await db.settings.create_index("key", unique=True)
    except Exception as e:
        logger.warning(f"Index creation skipped/partial: {e}")

async def seed_data():
    if not await db.admins.find_one({"username": "Aarav"}):
        pw_hash = bcrypt.hashpw("Krish@2026".encode(), bcrypt.gensalt()).decode()
        await db.admins.insert_one({"username": "Aarav", "password_hash": pw_hash})

    if await db.posts.count_documents({}) == 0:
        await db.posts.insert_many([{"id": str(uuid.uuid4()), **p} for p in DEFAULT_POSTS])

    if not await db.settings.find_one({"key": "election_open"}):
        await db.settings.insert_one({"key": "election_open", "value": "true"})

    if await db.users.count_documents({}) == 0:
        students = [
            {"admission_no": "SDPSS001", "name": "Aarav Sharma", "role": "student", "father_name": "Rajesh Sharma", "class_name": "XII-A", "subject": "", "designation": ""},
            {"admission_no": "SDPSS002", "name": "Ishita Verma", "role": "student", "father_name": "Mahesh Verma", "class_name": "XII-A", "subject": "", "designation": ""},
            {"admission_no": "SDPSS003", "name": "Krish Patel", "role": "student", "father_name": "Nikhil Patel", "class_name": "XII-B", "subject": "", "designation": ""},
            {"admission_no": "SDPSS004", "name": "Saanvi Gupta", "role": "student", "father_name": "Anil Gupta", "class_name": "XI-A", "subject": "", "designation": ""},
            {"admission_no": "SDPSS005", "name": "Vihaan Singh", "role": "student", "father_name": "Karan Singh", "class_name": "XI-B", "subject": "", "designation": ""},
            {"admission_no": "SDPSS006", "name": "Ananya Iyer", "role": "student", "father_name": "Suresh Iyer", "class_name": "X-A", "subject": "", "designation": ""},
            {"admission_no": "SDPSS007", "name": "Reyansh Mehta", "role": "student", "father_name": "Vivek Mehta", "class_name": "X-B", "subject": "", "designation": ""},
            {"admission_no": "SDPSS008", "name": "Diya Joshi", "role": "student", "father_name": "Manoj Joshi", "class_name": "IX-A", "subject": "", "designation": ""},
        ]
        teachers = [
            {"admission_no": "SDPSE01", "name": "Mrs. Anjali Rao", "role": "teacher", "father_name": "", "class_name": "", "subject": "Mathematics", "designation": "Sr. Teacher"},
            {"admission_no": "SDPSE02", "name": "Mr. Vikram Desai", "role": "teacher", "father_name": "", "class_name": "", "subject": "Physics", "designation": "HOD Science"},
            {"admission_no": "SDPSE03", "name": "Mrs. Pooja Saxena", "role": "teacher", "father_name": "", "class_name": "", "subject": "English", "designation": "Coordinator"},
        ]
        await db.users.insert_many(students + teachers)

    if await db.candidates.count_documents({}) == 0:
        photos = [
            "https://images.unsplash.com/photo-1693162274256-6bfe792b05e2?w=400&h=400&fit=crop",
            "https://images.unsplash.com/photo-1514960919797-5ff58c52e5ba?w=400&h=400&fit=crop",
            "https://images.unsplash.com/photo-1578390431312-f9ffd91de51b?w=400&h=400&fit=crop",
            "https://images.unsplash.com/photo-1596875422535-4267bf9495e0?w=400&h=400&fit=crop",
        ]
        symbols = ["Star", "Sun", "Book", "Tree"]
        seed_names = {
            "head_boy": ["Arjun Rao", "Kabir Khanna", "Dev Malhotra", "Yash Bhatt"],
            "head_girl": ["Riya Kapoor", "Meera Nair", "Tara Bose", "Nyra Shah"],
            "sports_skipper": ["Aditya Reddy", "Rohan Das", "Veer Sinha", "Ayaan Pillai"],
            "cultural_head": ["Aaradhya Jain", "Pari Mishra", "Zoya Ali", "Ira Chawla"],
            "discipline_head": ["Vivaan Bhatia", "Aarush Sethi", "Kian Roy", "Atharv Pandey"],
        }
        docs = []
        for pkey, names in seed_names.items():
            for i, nm in enumerate(names):
                docs.append({
                    "id": str(uuid.uuid4()), "post": pkey, "name": nm,
                    "photo": photos[i], "symbol": symbols[i], "symbol_image": "",
                })
        await db.candidates.insert_many(docs)


# ---------- Public Routes ----------
@api.get("/")
async def root():
    return {"message": "SDPS Election API"}

@api.get("/health")
async def health():
    return {"ok": True, "ts": datetime.now(timezone.utc).isoformat()}

@api.get("/posts")
async def list_posts():
    return await db.posts.find({}, {"_id": 0}).sort("order", 1).to_list(1000)

@api.get("/settings")
async def get_public_settings():
    docs = await db.settings.find({}, {"_id": 0}).to_list(100)
    return {d["key"]: d.get("value", "") for d in docs}

@api.get("/users/{admission_no}")
async def get_user(admission_no: str):
    adm = admission_no.strip()
    u = await db.users.find_one({"admission_no": adm}, {"_id": 0})
    if not u:
        raise HTTPException(status_code=404, detail="ID not found in records")
    voted = await db.votes.find_one({"admission_no": adm}, {"_id": 0, "id": 1})
    return {**u, "has_voted": bool(voted)}

@api.get("/candidates")
async def get_candidates(post: Optional[str] = None):
    q = {"post": post} if post else {}
    docs = await db.candidates.find(q, {"_id": 0}).to_list(1000)
    return [_lighten_candidate(c) for c in docs]

@api.get("/candidates/{cid}/photo")
async def get_candidate_photo(cid: str):
    return await _stream_candidate_image(cid, "photo")


@api.get("/candidates/{cid}/symbol")
async def get_candidate_symbol(cid: str):
    return await _stream_candidate_image(cid, "symbol_image")


async def _stream_candidate_image(cid: str, field: str):
    """Streams a (potentially huge base64) image only when needed.
    Browser caches it for a day, so subsequent kiosk views are instant."""
    c = await db.candidates.find_one({"id": cid}, {"_id": 0, field: 1})
    if not c:
        raise HTTPException(status_code=404, detail="Not found")
    src = c.get(field) or ""
    if not src:
        raise HTTPException(status_code=404, detail="No image")
    if src.startswith("data:"):
        try:
            header, b64 = src.split(",", 1)
            mime = header.split(";")[0].replace("data:", "") or "image/jpeg"
            import base64 as _b64
            raw = _b64.b64decode(b64)
            return StreamingResponse(
                io.BytesIO(raw),
                media_type=mime,
                headers={"Cache-Control": "public, max-age=86400, immutable"},
            )
        except Exception:
            raise HTTPException(status_code=500, detail="Invalid image data")
    from fastapi.responses import RedirectResponse
    return RedirectResponse(src, status_code=302)

@api.get("/bootstrap")
async def bootstrap():
    """Single round-trip: posts + all candidates + public settings.
    Drastically reduces first-load latency on Azure + Cosmos by avoiding
    N per-category fetches from the kiosk."""
    posts, cands, settings = await asyncio_gather_safe(
        db.posts.find({}, {"_id": 0}).sort("order", 1).to_list(1000),
        db.candidates.find({}, {"_id": 0}).to_list(2000),
        db.settings.find({}, {"_id": 0}).to_list(100),
    )
    return {
        "posts": posts,
        "candidates": [_lighten_candidate(c) for c in cands],
        "settings": {d["key"]: d.get("value", "") for d in settings},
    }

@api.post("/votes")
async def cast_vote(ballot: Ballot):
    s = await db.settings.find_one({"key": "election_open"}, {"_id": 0})
    if s and str(s.get("value", "true")).lower() in ("false", "0", "closed"):
        raise HTTPException(status_code=403, detail="Voting is currently closed")
    adm = ballot.admission_no.strip()
    user = await db.users.find_one({"admission_no": adm}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Voter not found")
    if await db.votes.find_one({"admission_no": adm}):
        raise HTTPException(status_code=400, detail="This ID has already cast their vote")
    keys = await active_post_keys()
    missing = [p for p in keys if p not in ballot.selections]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing votes for: {', '.join(missing)}")
    # Validate candidates with one bulk query
    cand_docs = await db.candidates.find(
        {"id": {"$in": list(ballot.selections.values())}},
        {"_id": 0, "id": 1, "post": 1},
    ).to_list(1000)
    cand_map = {c["id"]: c["post"] for c in cand_docs}
    for pkey, cid in ballot.selections.items():
        if pkey not in keys:
            raise HTTPException(status_code=400, detail=f"Unknown category: {pkey}")
        if cand_map.get(cid) != pkey:
            raise HTTPException(status_code=400, detail=f"Invalid candidate for {pkey}")
    doc = {
        "id": str(uuid.uuid4()), "admission_no": adm,
        "voter_name": user["name"], "voter_role": user.get("role", "student"),
        "voter_class": user.get("class_name", ""),
        "selections": ballot.selections, "timestamp": datetime.now(timezone.utc).isoformat(),
    }
    await db.votes.insert_one(doc)
    doc.pop("_id", None)
    return {"ok": True, "vote_id": doc["id"]}


# ---------- Admin Auth ----------
@api.post("/admin/login")
async def admin_login(body: AdminLogin):
    user = await db.admins.find_one({"username": body.username})
    if not user or not bcrypt.checkpw(body.password.encode(), user["password_hash"].encode()):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    return {"token": make_token(body.username), "username": body.username}

@api.get("/admin/me")
async def admin_me(_: str = Depends(verify_admin)):
    return {"ok": True, "username": _}


# ---------- Admin: Users ----------
@api.get("/admin/users")
async def list_users(role: Optional[str] = None, _: str = Depends(verify_admin)):
    q = {"role": role} if role else {}
    docs = await db.users.find(q, {"_id": 0}).to_list(20000)
    voted_docs = await db.votes.find({}, {"_id": 0, "admission_no": 1}).to_list(30000)
    voted_set = {v["admission_no"] for v in voted_docs}
    for d in docs:
        d["has_voted"] = d["admission_no"] in voted_set
    return docs

@api.post("/admin/users/upload")
async def upload_users(role: str, file: UploadFile = File(...), _: str = Depends(verify_admin)):
    if role not in ("student", "teacher"):
        raise HTTPException(status_code=400, detail="role must be 'student' or 'teacher'")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")
    if not rows:
        raise HTTPException(status_code=400, detail="Empty file")
    header = [str(c).strip().lower() if c else "" for c in rows[0]]

    def idx(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return -1

    i_adm = idx("admission_no", "admission no", "admission number", "id")
    i_name = idx("name", "full name", "student name", "teacher name")
    if i_adm < 0 or i_name < 0:
        raise HTTPException(status_code=400, detail="Header must include: admission_no, name (and role-specific columns)")

    if role == "student":
        i_father = idx("father_name", "father name", "father's name", "fathers name", "father")
        i_class = idx("class_name", "class", "class name")
        if i_father < 0:
            raise HTTPException(status_code=400, detail="Student file requires column: father_name")
        i_subject = -1
        i_desig = -1
    else:
        i_subject = idx("subject", "subjects")
        i_desig = idx("designation", "role", "post")
        i_father = -1
        i_class = -1

    inserted, updated = 0, 0
    for r in rows[1:]:
        if not r or i_adm >= len(r) or not r[i_adm]:
            continue
        adm = str(r[i_adm]).strip()
        doc = {
            "admission_no": adm,
            "name": str(r[i_name]).strip() if i_name >= 0 and r[i_name] else "",
            "role": role,
            "father_name": str(r[i_father]).strip() if i_father >= 0 and r[i_father] else "",
            "class_name": str(r[i_class]).strip() if i_class >= 0 and r[i_class] else "",
            "subject": str(r[i_subject]).strip() if i_subject >= 0 and r[i_subject] else "",
            "designation": str(r[i_desig]).strip() if i_desig >= 0 and r[i_desig] else "",
        }
        res = await db.users.update_one({"admission_no": adm}, {"$set": doc}, upsert=True)
        if res.upserted_id:
            inserted += 1
        else:
            updated += 1
    return {"inserted": inserted, "updated": updated}

@api.delete("/admin/users/{admission_no}")
async def delete_user(admission_no: str, _: str = Depends(verify_admin)):
    res = await db.users.delete_one({"admission_no": admission_no})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}

@api.get("/admin/template/{role}")
async def download_template(role: str, _: str = Depends(verify_admin)):
    if role not in ("student", "teacher"):
        raise HTTPException(status_code=400, detail="Invalid role")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students" if role == "student" else "Teachers"
    if role == "student":
        headers = ["admission_no", "name", "father_name", "class_name"]
        sample = [
            ["SDPSS001", "Aarav Sharma", "Rajesh Sharma", "XII-A"],
            ["SDPSS002", "Ishita Verma", "Mahesh Verma", "XII-A"],
            ["SDPSS003", "Krish Patel", "Nikhil Patel", "XII-B"],
        ]
    else:
        headers = ["admission_no", "name", "subject", "designation"]
        sample = [
            ["SDPSE01", "Mrs. Anjali Rao", "Mathematics", "Sr. Teacher"],
            ["SDPSE02", "Mr. Vikram Desai", "Physics", "HOD Science"],
            ["SDPSE03", "Mrs. Pooja Saxena", "English", "Coordinator"],
        ]
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="0F3C8A")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20
    for r, row in enumerate(sample, 2):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"sdps_{role}_template.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------- Admin: Posts ----------
@api.get("/admin/posts")
async def admin_posts(_: str = Depends(verify_admin)):
    docs = await db.posts.find({}, {"_id": 0}).sort("order", 1).to_list(1000)
    cands = await db.candidates.find({}, {"_id": 0, "post": 1}).to_list(2000)
    counts = {}
    for c in cands:
        counts[c["post"]] = counts.get(c["post"], 0) + 1
    vote_docs = await db.votes.find({}, {"_id": 0, "selections": 1}).to_list(30000)
    votes_by_post = {}
    for v in vote_docs:
        for k in v.get("selections", {}):
            votes_by_post[k] = votes_by_post.get(k, 0) + 1
    for d in docs:
        d["candidate_count"] = counts.get(d["key"], 0)
        d["vote_count"] = votes_by_post.get(d["key"], 0)
    return docs

@api.post("/admin/posts")
async def create_post(body: PostCreate, _: str = Depends(verify_admin)):
    title = body.title.strip()
    if not title:
        raise HTTPException(status_code=400, detail="Title required")
    base = slugify(title)
    key = base
    n = 1
    while await db.posts.find_one({"key": key}):
        n += 1
        key = f"{base}_{n}"
    if body.order is None:
        last = await db.posts.find().sort("order", -1).limit(1).to_list(1)
        order = (last[0]["order"] + 1) if last else 1
    else:
        order = body.order
    doc = {"id": str(uuid.uuid4()), "key": key, "title": title, "order": order}
    await db.posts.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/admin/posts/{pid}")
async def update_post(pid: str, body: PostUpdate, _: str = Depends(verify_admin)):
    upd = {k: v for k, v in body.model_dump().items() if v is not None}
    if not upd:
        raise HTTPException(status_code=400, detail="Nothing to update")
    res = await db.posts.update_one({"id": pid}, {"$set": upd})
    if not res.matched_count:
        raise HTTPException(status_code=404, detail="Not found")
    return await db.posts.find_one({"id": pid}, {"_id": 0})

@api.delete("/admin/posts/{pid}")
async def delete_post(pid: str, _: str = Depends(verify_admin)):
    p = await db.posts.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Not found")
    has_votes = False
    vote_docs = await db.votes.find({}, {"_id": 0, "selections": 1}).to_list(30000)
    for v in vote_docs:
        if p["key"] in (v.get("selections") or {}):
            has_votes = True
            break
    if has_votes:
        raise HTTPException(status_code=400, detail="Cannot delete: votes exist for this category. Reset votes first.")
    await db.candidates.delete_many({"post": p["key"]})
    await db.posts.delete_one({"id": pid})
    return {"ok": True}


# ---------- Admin: Candidates ----------
@api.post("/admin/candidates")
async def create_candidate(body: CandidateCreate, _: str = Depends(verify_admin)):
    keys = await active_post_keys()
    if body.post not in keys:
        raise HTTPException(status_code=400, detail="Invalid category")
    cand = Candidate(**body.model_dump())
    await db.candidates.insert_one(cand.model_dump())
    return _lighten_candidate(cand.model_dump())

@api.put("/admin/candidates/{cid}")
async def update_candidate(cid: str, body: CandidateUpdate, _: str = Depends(verify_admin)):
    upd = {k: v for k, v in body.model_dump().items() if v is not None}
    if "post" in upd:
        keys = await active_post_keys()
        if upd["post"] not in keys:
            raise HTTPException(status_code=400, detail="Invalid category")
    res = await db.candidates.update_one({"id": cid}, {"$set": upd})
    if not res.matched_count:
        raise HTTPException(status_code=404, detail="Not found")
    doc = await db.candidates.find_one({"id": cid}, {"_id": 0})
    return _lighten_candidate(doc) if doc else None

@api.delete("/admin/candidates/{cid}")
async def delete_candidate(cid: str, _: str = Depends(verify_admin)):
    res = await db.candidates.delete_one({"id": cid})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


# ---------- Public Notice Board (turnout only, no leaders) ----------
@api.get("/board")
async def public_board():
    posts = await db.posts.find({}, {"_id": 0}).sort("order", 1).to_list(1000)
    votes = await db.votes.find({}, {"_id": 0, "admission_no": 1, "voter_role": 1, "voter_class": 1, "timestamp": 1}).to_list(30000)
    total_users = await db.users.count_documents({})
    total_students = await db.users.count_documents({"role": "student"})
    total_teachers = await db.users.count_documents({"role": "teacher"})
    voted_set = {v["admission_no"] for v in votes}

    class_groups: Dict[str, Dict[str, int]] = {}
    student_docs = await db.users.find({"role": "student"}, {"_id": 0, "admission_no": 1, "class_name": 1}).to_list(20000)
    for u in student_docs:
        cls = u.get("class_name") or "Unassigned"
        g = class_groups.setdefault(cls, {"class_name": cls, "total": 0, "voted": 0})
        g["total"] += 1
        if u["admission_no"] in voted_set:
            g["voted"] += 1
    class_breakdown = sorted(class_groups.values(), key=lambda x: x["class_name"])

    voted_students = sum(1 for v in votes if v.get("voter_role") == "student")
    voted_teachers = sum(1 for v in votes if v.get("voter_role") == "teacher")

    s = await db.settings.find_one({"key": "election_open"}, {"_id": 0})
    election_open = str((s or {}).get("value", "true")).lower() != "false"
    last_ts = max((v.get("timestamp", "") for v in votes), default="")

    return {
        "election_open": election_open,
        "categories_count": len(posts),
        "total_users": total_users,
        "total_students": total_students,
        "total_teachers": total_teachers,
        "total_voted": len(votes),
        "voted_students": voted_students,
        "voted_teachers": voted_teachers,
        "pending": max(0, total_users - len(votes)),
        "turnout_pct": round((len(votes) / total_users * 100), 1) if total_users else 0,
        "class_breakdown": class_breakdown,
        "last_vote_at": last_ts,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }


# ---------- Admin-protected results ----------
@api.get("/results")
async def public_results(_: str = Depends(verify_admin)):
    posts = await db.posts.find({}, {"_id": 0}).sort("order", 1).to_list(1000)
    candidates = await db.candidates.find({}, {"_id": 0}).to_list(2000)
    votes = await db.votes.find({}, {"_id": 0, "selections": 1, "voter_class": 1, "admission_no": 1}).to_list(30000)
    total_users = await db.users.count_documents({})

    counts = {}
    for v in votes:
        for cid in (v.get("selections") or {}).values():
            counts[cid] = counts.get(cid, 0) + 1

    by_post = {p["key"]: [] for p in posts}
    for c in candidates:
        adj = int(c.get("adjustment") or 0)
        photo = c.get("photo", "")
        if photo.startswith("data:"):
            photo = f"{PUBLIC_BACKEND_URL}/api/candidates/{c['id']}/photo"
        entry = {
            "candidate_id": c["id"], "name": c["name"],
            "photo": photo, "symbol": c.get("symbol", ""),
            "votes": counts.get(c["id"], 0) + adj,
        }
        if c["post"] in by_post:
            by_post[c["post"]].append(entry)
    for k in by_post:
        by_post[k].sort(key=lambda x: x["votes"], reverse=True)

    return {
        "posts": posts,
        "by_post": by_post,
        "winners": {p["key"]: (by_post[p["key"]][0] if by_post[p["key"]] else None) for p in posts},
        "total_voted": len(votes),
        "total_users": total_users,
        "turnout_pct": round((len(votes) / total_users * 100), 1) if total_users else 0,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }


# ---------- Admin: Stats ----------
@api.get("/admin/stats")
async def stats(_: str = Depends(verify_admin)):
    posts = await db.posts.find({}, {"_id": 0}).sort("order", 1).to_list(1000)
    post_keys = [p["key"] for p in posts]
    candidates = await db.candidates.find({}, {"_id": 0}).to_list(2000)
    cand_map = {c["id"]: c for c in candidates}
    votes = await db.votes.find({}, {"_id": 0}).to_list(30000)
    total_users = await db.users.count_documents({})
    students_total = await db.users.count_documents({"role": "student"})
    teachers_total = await db.users.count_documents({"role": "teacher"})

    by_post = {p: [] for p in post_keys}
    counts = {}
    for v in votes:
        for pkey, cid in v.get("selections", {}).items():
            counts[cid] = counts.get(cid, 0) + 1
    for c in candidates:
        adj = int(c.get("adjustment") or 0)
        real = counts.get(c["id"], 0)
        photo = c.get("photo", "")
        if photo.startswith("data:"):
            photo = f"{PUBLIC_BACKEND_URL}/api/candidates/{c['id']}/photo"
        entry = {
            "candidate_id": c["id"], "name": c["name"],
            "photo": photo, "symbol": c.get("symbol", ""),
            "real_votes": real, "adjustment": adj,
            "votes": real + adj,
        }
        if c["post"] in by_post:
            by_post[c["post"]].append(entry)
    winners = {}
    for pkey, lst in by_post.items():
        lst.sort(key=lambda x: x["votes"], reverse=True)
        winners[pkey] = lst[0] if lst else None

    class_groups: Dict[str, Dict[str, int]] = {}
    voted_set = {v["admission_no"] for v in votes}
    student_docs = await db.users.find({"role": "student"}, {"_id": 0, "admission_no": 1, "class_name": 1}).to_list(20000)
    for u in student_docs:
        cls = u.get("class_name") or "Unassigned"
        g = class_groups.setdefault(cls, {"class_name": cls, "total": 0, "voted": 0})
        g["total"] += 1
        if u["admission_no"] in voted_set:
            g["voted"] += 1
    class_breakdown = sorted(class_groups.values(), key=lambda x: x["class_name"])

    return {
        "posts": posts,
        "total_users": total_users,
        "total_students": students_total,
        "total_teachers": teachers_total,
        "total_voted": len(votes),
        "turnout_pct": round((len(votes) / total_users * 100), 1) if total_users else 0,
        "by_post": by_post,
        "winners": winners,
        "class_breakdown": class_breakdown,
        "votes": [
            {
                "id": v.get("id"),
                "admission_no": v["admission_no"],
                "voter_name": v.get("voter_name", v.get("student_name", "")),
                "voter_role": v.get("voter_role", "student"),
                "voter_class": v.get("voter_class", ""),
                "timestamp": v.get("timestamp", ""),
                "selections": v.get("selections", {}),
                "selection_names": {
                    pkey: cand_map.get(cid, {}).get("name", "Unknown")
                    for pkey, cid in v.get("selections", {}).items()
                },
            } for v in votes
        ],
    }


# ---------- Admin: Vote Manipulation ----------
@api.put("/admin/votes/{vote_id}")
async def edit_vote(vote_id: str, body: BallotUpdate, _: str = Depends(verify_admin)):
    v = await db.votes.find_one({"id": vote_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Ballot not found")
    keys = await active_post_keys()
    cand_docs = await db.candidates.find(
        {"id": {"$in": list(body.selections.values())}},
        {"_id": 0, "id": 1, "post": 1},
    ).to_list(1000)
    cand_map = {c["id"]: c["post"] for c in cand_docs}
    for pkey, cid in body.selections.items():
        if pkey not in keys:
            raise HTTPException(status_code=400, detail=f"Unknown category: {pkey}")
        if cand_map.get(cid) != pkey:
            raise HTTPException(status_code=400, detail=f"Invalid candidate for {pkey}")
    await db.votes.update_one({"id": vote_id}, {"$set": {"selections": body.selections}})
    return {"ok": True}

@api.delete("/admin/votes/{vote_id}")
async def delete_vote(vote_id: str, _: str = Depends(verify_admin)):
    res = await db.votes.delete_one({"id": vote_id})
    if not res.deleted_count:
        raise HTTPException(status_code=404, detail="Ballot not found")
    return {"ok": True}


# ---------- Admin: Settings ----------
@api.get("/admin/settings")
async def admin_settings(_: str = Depends(verify_admin)):
    docs = await db.settings.find({}, {"_id": 0}).to_list(100)
    return {d["key"]: d.get("value", "") for d in docs}

@api.put("/admin/settings/{key}")
async def update_setting(key: str, body: SettingValue, _: str = Depends(verify_admin)):
    await db.settings.update_one({"key": key}, {"$set": {"key": key, "value": body.value}}, upsert=True)
    return {"ok": True, "key": key}


# ---------- Admin: Reset ----------
@api.post("/admin/reset/votes")
async def reset_votes(_: str = Depends(verify_admin)):
    res = await db.votes.delete_many({})
    return {"ok": True, "deleted_votes": res.deleted_count}

@api.post("/admin/reset/all")
async def reset_all(_: str = Depends(verify_admin)):
    v = await db.votes.delete_many({})
    c = await db.candidates.delete_many({})
    u = await db.users.delete_many({})
    return {
        "ok": True,
        "deleted_votes": v.deleted_count,
        "deleted_candidates": c.deleted_count,
        "deleted_users": u.deleted_count,
    }


# ---------- asyncio helper (kept simple, no extra import at top) ----------
async def asyncio_gather_safe(*coros):
    import asyncio as _asyncio
    return await _asyncio.gather(*coros)


# ---------- Wire up app ----------
app.include_router(api)

cors_origins_raw = os.environ.get(
    'CORS_ORIGINS',
    'https://sdps-election-web.vercel.app,http://localhost:3000',
)
cors_origins = [o.strip() for o in cors_origins_raw.split(',') if o.strip()] or ['*']
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=cors_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup():
    await ensure_indexes()
    await seed_data()
    logger.info("SDPS Election API started.")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
